#!/usr/bin/env python3
"""Network Accessibility and Robustness.

Plan: Summarize walking, motorized, mixed-mode, capacity, and facility-loss evidence.
Framework: Sections 5-7 interpret walking and motorized results as accessibility
bounds, capacity allocation as a conditional stress test, and facility loss as
modeled sensitivity rather than observed denial of shelter.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
ACCESS_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/walking_and_motorized_accessibility_summary.csv"
MIXED_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/mixed_mode_accessibility_summary.csv"
MUNICIPALITY_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/municipality_mixed_mode_accessibility.csv"
CAPACITY_PATH = ROOT / "data/exp/primary-capacity-constrained-allocation/capacity_threshold_sensitivity.csv"
RANDOM_PATH = ROOT / "data/exp/shelter-robustness/facility_unavailability_random_summary.csv"
FACILITY_PATH = ROOT / "data/exp/shelter-robustness/facility_unavailability_sensitivity.csv"
CRITICAL_PATH = ROOT / "data/exp/shelter-robustness/critical_single_shelter_loss.csv"
OUTPUT_PATH = ROOT / "data/exp/legacy-outputs/Table_network_adequacy_and_robustness.xlsx"
SHEET_NAME = "Network Robustness"

DEMAND_MEASURE = "Observed-Use Stress Demand High Housing-Loss Weighted"
ENGLISH_NAMES = {"43100": "Kumamoto City", "43202": "Yatsushiro", "43213": "Uki"}
CRITICAL_NAMES = {
    "E4321300034111": "Toyofuku Elementary School Gymnasium (Uki)",
    "E4321300010111": "Ogawa Disaster Prevention Base Center (Uki)",
    "E4321300011111": "Rapport Cultural Center (Uki)",
    "E4320200016111": "Kagami Elementary School (Yatsushiro)",
}

COLUMNS = [
    "Evidence Block", "Scenario", "Vehicle-Enabled Demand Share (%)",
    "Road Speed Factor", "Time Threshold (min)", "Capacity per Shelter",
    "Maximum Open Shelters", "Available Shelters", "Accessible or Assigned Demand",
    "Accessible or Assigned Share (%)", "Explanation Gap or Service Loss",
    "Solution Qualification",
]


def empty_row() -> dict[str, object]:
    return {column: pd.NA for column in COLUMNS}


def build_table() -> pd.DataFrame:
    access = pd.read_csv(ACCESS_PATH)
    mixed = pd.read_csv(MIXED_PATH)
    capacity = pd.read_csv(CAPACITY_PATH)
    random = pd.read_csv(RANDOM_PATH)
    facility = pd.read_csv(FACILITY_PATH)
    critical = pd.read_csv(CRITICAL_PATH)
    municipality = pd.read_csv(MUNICIPALITY_PATH, dtype={"Municipality Code": str})
    baseline = facility.loc[facility["Failure Mode"].eq("baseline")].iloc[0]
    baseline_served = float(baseline["Maximum Served Demand"])
    total = float(baseline["Scenario Demand"])
    rows: list[dict[str, object]] = []

    walk = access.loc[(access["Mode"].eq("Walking")) & (access["Demand Measure"].eq(DEMAND_MEASURE)) & (access["Time Threshold (min)"].eq(15))].iloc[0]
    row = empty_row(); row.update({
        "Evidence Block": "Accessibility bounds", "Scenario": "Walking bound",
        "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
        "Accessible or Assigned Demand": round(float(walk["Accessible Demand"])),
        "Accessible or Assigned Share (%)": float(walk["Accessible Percent"]),
        "Explanation Gap or Service Loss": round(total - float(walk["Accessible Demand"])),
        "Solution Qualification": "4 km/h walking network",
    }); rows.append(row)
    for speed_factor in (0.25, 0.50, 1.00):
        motor = access.loc[(access["Mode"].eq("Motorized")) & (access["Demand Measure"].eq(DEMAND_MEASURE)) & (access["Road Speed Factor"].eq(speed_factor)) & (access["Time Threshold (min)"].eq(15))].iloc[0]
        row = empty_row(); row.update({
            "Evidence Block": "Accessibility bounds", "Scenario": "Motorized bound",
            "Vehicle-Enabled Demand Share (%)": 100, "Road Speed Factor": speed_factor,
            "Time Threshold (min)": 15,
            "Accessible or Assigned Demand": round(float(motor["Accessible Demand"])),
            "Accessible or Assigned Share (%)": float(motor["Accessible Percent"]),
            "Explanation Gap or Service Loss": round(total - float(motor["Accessible Demand"])),
            "Solution Qualification": "Connectors walked; road travel motorized",
        }); rows.append(row)
    for share in (0.25, 0.50, 0.75, 1.00):
        source = mixed.loc[mixed["Vehicle-Enabled Demand Share"].eq(share)].iloc[0]
        row = empty_row(); row.update({
            "Evidence Block": "Accessibility bounds", "Scenario": "Mixed-mode sensitivity",
            "Vehicle-Enabled Demand Share (%)": 100 * share,
            "Road Speed Factor": float(source["Motorized Road Speed Factor"]),
            "Time Threshold (min)": int(source["Time Threshold (min)"]),
            "Accessible or Assigned Demand": round(float(source["Accessible Demand"])),
            "Accessible or Assigned Share (%)": float(source["Accessible Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Explanation Gap"])),
            "Solution Qualification": "Share is a sensitivity parameter",
        }); rows.append(row)

    for capacity_value in (25, 50, 100, 200):
        source = capacity.loc[(capacity["Capacity per Open Shelter"].eq(capacity_value)) & (capacity["Maximum Open Shelters"].eq(415))].iloc[0]
        role = "Stress" if capacity_value == 50 else "Central" if capacity_value == 100 else "Sensitivity"
        row = empty_row(); row.update({
            "Evidence Block": "Capacity allocation", "Scenario": f"{role} capacity case",
            "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
            "Capacity per Shelter": capacity_value, "Maximum Open Shelters": 415,
            "Available Shelters": 1156,
            "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
            "Accessible or Assigned Share (%)": float(source["Served Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Unmet Demand"])),
            "Solution Qualification": "Optimal" if bool(source["Proven Optimal"]) else "Time-limit incumbent; lower bound",
        }); rows.append(row)

    row = empty_row(); row.update({
        "Evidence Block": "Facility unavailability", "Scenario": "Baseline allocation",
        "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
        "Capacity per Shelter": 50, "Maximum Open Shelters": 415,
        "Available Shelters": int(baseline["Available Shelters"]),
        "Accessible or Assigned Demand": round(baseline_served),
        "Accessible or Assigned Share (%)": float(baseline["Served Percent"]),
        "Explanation Gap or Service Loss": 0, "Solution Qualification": "Optimal",
    }); rows.append(row)
    for share in (0.10, 0.20, 0.30):
        source = random.loc[random["Unavailability Share"].eq(share)].iloc[0]
        row = empty_row(); row.update({
            "Evidence Block": "Facility unavailability", "Scenario": f"Random {int(share * 100)}% removal",
            "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
            "Capacity per Shelter": 50, "Maximum Open Shelters": 415,
            "Available Shelters": round(1156 * (1 - share)),
            "Accessible or Assigned Demand": round(float(source["Mean_Served_Demand"])),
            "Accessible or Assigned Share (%)": float(source["Mean_Served_Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Mean_Service_Loss"])),
            "Solution Qualification": f"Mean of {int(source['Draws'])} reproducible draws",
        }); rows.append(row)
    targeted = facility.loc[facility["Failure Mode"].eq("targeted_high_reachable_pressure")]
    for _, source in targeted.sort_values("Unavailability Share").iterrows():
        share = float(source["Unavailability Share"])
        row = empty_row(); row.update({
            "Evidence Block": "Facility unavailability", "Scenario": f"Pressure-targeted {int(share * 100)}% removal",
            "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
            "Capacity per Shelter": 50, "Maximum Open Shelters": 415,
            "Available Shelters": int(source["Available Shelters"]),
            "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
            "Accessible or Assigned Share (%)": float(source["Served Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Service Loss from Baseline"])),
            "Solution Qualification": "Optimal",
        }); rows.append(row)

    walk_m = municipality.loc[municipality["Vehicle-Enabled Demand Share"].eq(0)].set_index("Municipality Code")
    motor_m = municipality.loc[municipality["Vehicle-Enabled Demand Share"].eq(1)].set_index("Municipality Code")
    gains = walk_m[["Stress Load", "Accessible Demand", "Accessible Percent"]].join(motor_m[["Accessible Demand", "Accessible Percent"]], lsuffix=" Walk", rsuffix=" Motor")
    gains["Demand Gain"] = gains["Accessible Demand Motor"] - gains["Accessible Demand Walk"]
    for code, source in gains.sort_values("Demand Gain", ascending=False).head(3).iterrows():
        row = empty_row(); row.update({
            "Evidence Block": "Municipality mode gap", "Scenario": ENGLISH_NAMES[str(code)],
            "Vehicle-Enabled Demand Share (%)": 100, "Road Speed Factor": 0.5,
            "Time Threshold (min)": 15,
            "Accessible or Assigned Demand": round(float(source["Accessible Demand Motor"])),
            "Accessible or Assigned Share (%)": float(source["Accessible Percent Motor"]),
            "Explanation Gap or Service Loss": round(float(source["Demand Gain"])),
            "Solution Qualification": f"Gain over walking bound; walk {source['Accessible Percent Walk']:.1f}%",
        }); rows.append(row)

    for _, source in critical.sort_values("Single-Shelter Service-Loss Lower Bound", ascending=False).head(4).iterrows():
        shelter_id = str(source["Shelter ID"])
        row = empty_row(); row.update({
            "Evidence Block": "Single-shelter loss", "Scenario": CRITICAL_NAMES[shelter_id],
            "Vehicle-Enabled Demand Share (%)": 0, "Time Threshold (min)": 15,
            "Capacity per Shelter": 50, "Maximum Open Shelters": 415,
            "Available Shelters": 1155,
            "Accessible or Assigned Demand": round(float(source["Served Demand after Removal"])),
            "Accessible or Assigned Share (%)": 100 * float(source["Served Demand after Removal"]) / total,
            "Explanation Gap or Service Loss": round(float(source["Single-Shelter Service-Loss Lower Bound"])),
            "Solution Qualification": "Screened among 30 high-pressure shelters",
        }); rows.append(row)

    table = pd.DataFrame(rows, columns=COLUMNS)
    if table.shape != (26, 12):
        raise RuntimeError(f"Expected a 26 x 12 table, found {table.shape}.")
    if not table["Scenario"].eq("Random 20% removal").any():
        raise RuntimeError("The random 20% facility-removal result is missing.")
    return table


def notes_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Demand surface", "All rows use the 10,467-scaled high-housing-loss stress surface; it is counterfactual spatial demand, not observed municipality shelter use."),
            ("Accessibility bounds", "Walking is a restrictive lower bound. Motorized travel uses assumed road speeds and walking connectors; mixed-mode shares are sensitivity parameters."),
            ("Central motorized assumption", "The central motorized comparison applies 50% of baseline road speed within 15 minutes."),
            ("Capacity roles", "The 100-person case is central, 50 persons is a conservative stress case, and 25 and 200 persons are sensitivity cases."),
            ("Explanation gap", "The difference between total stress load and modeled accessible or assigned demand is not observed refusal or unsheltered population."),
            ("Random unavailability", "Each random-removal row reports the mean of 30 reproducible draws; the 20% result is included."),
            ("Targeted unavailability", "Pressure-targeted removal orders shelters by reachable stress pressure and is an adverse sensitivity test."),
            ("Single-shelter screen", "Losses cover the 30 highest reachable-pressure shelters and therefore do not constitute an exhaustive ranking of all 1,156 general shelters."),
            ("Solution status", "The 25-person, 415-opening capacity case is a time-limit incumbent and a lower bound on served demand; other displayed allocation cases are optimal or near-optimal as noted."),
            ("Primary interpretation", "The accessibility ceiling changes far more than assigned share across plausible capacity cases, indicating that accessibility is the binding planning dimension."),
        ],
        columns=["Note", "Definition or Limitation"],
    )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    thin = Border(bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.2, color="172033")
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:L{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 75
    worksheet.print_area = f"A1:L{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.18, right=0.18, top=0.25, bottom=0.25, header=0.10, footer=0.10)
    for cell in worksheet[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 58
    fills = {"Accessibility bounds": "DDEBF7", "Capacity allocation": "FFF1D6", "Facility unavailability": "FDE7E3", "Municipality mode gap": "E8F3EC", "Single-shelter loss": "F2EBF6"}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font; cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = thin
        row[0].fill = PatternFill("solid", fgColor=fills[str(row[0].value)])
        row[0].font = Font(name="Aptos", size=8.2, bold=True, color="17365D")
        for index in range(2, 11):
            row[index].alignment = Alignment(horizontal="right", vertical="center")
        for index in (2, 4, 9): row[index].number_format = "0.0"
        row[3].number_format = "0.00"
        for index in (5, 6, 7, 8, 10): row[index].number_format = "#,##0"
        worksheet.row_dimensions[row[0].row].height = 37
    for column, width in {"A": 25, "B": 43, "C": 20, "D": 17, "E": 18, "F": 18, "G": 19, "H": 17, "I": 25, "J": 24, "K": 24, "L": 43}.items():
        worksheet.column_dimensions[column].width = width
    excel_table = Table(displayName="NetworkAccessibilityRobustness", ref=f"A1:L{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(excel_table)
    notes = workbook["Notes"]
    notes.sheet_view.showGridLines = False; notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 28; notes.column_dimensions["B"].width = 112
    for cell in notes[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font; cell.alignment = Alignment(vertical="top", wrap_text=True); cell.border = thin
        row[0].font = Font(name="Aptos", size=8.2, bold=True, color="17365D")
        notes.row_dimensions[row[0].row].height = 40
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 27 or worksheet.max_column != 12:
        raise RuntimeError(f"Unexpected main-sheet dimensions: {worksheet.max_row} x {worksheet.max_column}.")
    if worksheet.merged_cells.ranges:
        raise RuntimeError("Merged cells are not permitted in article-facing tables.")
    if workbook.sheetnames != [SHEET_NAME, "Notes"]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        build_table().to_excel(writer, sheet_name=SHEET_NAME, index=False)
        notes_table().to_excel(writer, sheet_name="Notes", index=False)
    style_workbook(OUTPUT_PATH); verify_workbook(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
