#!/usr/bin/env python3
"""Render a styled worksheet to a PNG review copy using openpyxl and Pillow."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


# Hiragino Sans GB covers Latin, Japanese kana, and the CJK ideographs used in
# Kumamoto municipality names, avoiding missing-glyph boxes in review images.
FONT_REGULAR = Path("/System/Library/Fonts/Hiragino Sans GB.ttc")
FONT_BOLD = Path("/System/Library/Fonts/Hiragino Sans GB.ttc")


def excel_color(cell: object, default: str = "FFFFFF") -> str:
    fill = cell.fill
    if fill.fill_type != "solid":
        return default
    value = fill.fgColor.rgb
    if value and len(value) >= 6:
        return value[-6:]
    return default


def text_color(cell: object, default: str = "172033") -> str:
    value = cell.font.color
    if value is not None and value.type == "rgb" and value.rgb:
        return value.rgb[-6:]
    return default


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    path = FONT_BOLD if bold else FONT_REGULAR
    return ImageFont.truetype(str(path), size=size)


def formatted_value(cell: object) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        number_format = str(cell.number_format)
        if "%" in number_format:
            decimals = 1 if ".0" in number_format else 0
            return f"{value * 100:.{decimals}f}%"
        if "0.00" in number_format:
            return f"{value:,.2f}"
        if "0.0" in number_format:
            return f"{value:,.1f}"
        if "0" in number_format:
            return f"{value:,.0f}"
    return str(value)


def wrap_text(draw: ImageDraw.ImageDraw, text: str, text_font: ImageFont.FreeTypeFont, width: int) -> list[str]:
    if not text:
        return []
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        trial = word if not current else f"{current} {word}"
        if draw.textlength(trial, font=text_font) <= width or not current:
            current = trial
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--start-data-row", type=int)
    parser.add_argument("--end-data-row", type=int)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active
    scale = 6.5
    column_widths = [
        max(70, int((worksheet.column_dimensions[cell.column_letter].width or 12) * scale))
        for cell in worksheet[2]
    ]
    if args.start_data_row or args.end_data_row:
        start = args.start_data_row or 3
        end = args.end_data_row or worksheet.max_row
        row_numbers = [1, 2, *range(start, end + 1)]
    else:
        row_numbers = list(range(1, worksheet.max_row + 1))
    row_heights = [
        max(30, int((worksheet.row_dimensions[row].height or 22) * 1.65))
        for row in row_numbers
    ]
    first_row_is_merged_title = any(
        merged.min_row == 1
        and merged.max_row == 1
        and merged.min_col == 1
        and merged.max_col == worksheet.max_column
        for merged in worksheet.merged_cells.ranges
    )
    image = Image.new("RGB", (sum(column_widths) + 2, sum(row_heights) + 2), "#FFFFFF")
    draw = ImageDraw.Draw(image)

    y = 1
    for row_number, row_height in zip(row_numbers, row_heights):
        if row_number == 1 and first_row_is_merged_title:
            cell = worksheet.cell(row=1, column=1)
            row_width = sum(column_widths)
            background = excel_color(cell)
            draw.rectangle(
                (1, y, 1 + row_width, y + row_height),
                fill=f"#{background}",
                outline="#D0D5DD",
                width=1,
            )
            text_font = font(23, True)
            draw.text(
                (8, y + max(4, (row_height - 28) // 2)),
                formatted_value(cell),
                fill=f"#{text_color(cell)}",
                font=text_font,
            )
            y += row_height
            continue
        x = 1
        for column_number, column_width in enumerate(column_widths, start=1):
            cell = worksheet.cell(row=row_number, column=column_number)
            background = excel_color(cell)
            draw.rectangle(
                (x, y, x + column_width, y + row_height),
                fill=f"#{background}",
                outline="#D0D5DD",
                width=1,
            )
            value = formatted_value(cell)
            is_header = row_number == (2 if first_row_is_merged_title else 1)
            size = 16 if is_header else 14
            text_font = font(size, bool(cell.font.bold) or is_header)
            lines = wrap_text(draw, value, text_font, max(15, column_width - 12))
            line_height = int(size * 1.22)
            total_height = len(lines) * line_height
            if is_header or (first_row_is_merged_title and row_number == 1):
                text_y = y + max(5, (row_height - total_height) // 2)
            else:
                text_y = y + 6
            alignment = cell.alignment.horizontal
            for line in lines:
                line_width = draw.textlength(line, font=text_font)
                if alignment == "right":
                    text_x = x + column_width - line_width - 6
                elif alignment == "center" or is_header:
                    text_x = x + (column_width - line_width) / 2
                else:
                    text_x = x + 6
                draw.text((text_x, text_y), line, fill=f"#{text_color(cell)}", font=text_font)
                text_y += line_height
            x += column_width
        y += row_height

    args.output.parent.mkdir(parents=True, exist_ok=True)
    image.save(args.output, format="PNG", optimize=True)
    print(f"Saved {args.output}")


if __name__ == "__main__":
    main()
