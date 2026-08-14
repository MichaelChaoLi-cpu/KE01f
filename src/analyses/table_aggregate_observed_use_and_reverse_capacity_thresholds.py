#!/usr/bin/env python3
"""Aggregate Stress Load and Reverse Capacity Thresholds.

Plan: Compare four official event-use snapshots with standardized capacity cases,
then report reverse critical capacities for four spatial stress surfaces.
Framework: Sections 5-7 separate prefecture arithmetic from municipality-contained
ceilings and treat 10,467 as an observed aggregate used to scale counterfactual
spatial stress loads, not as observed unmet demand.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
SNAPSHOT_PATH = ROOT / "data/exp/capacity-threshold-estimate/event_snapshot_capacity_thresholds.csv"
CRITICAL_PATH = ROOT / "data/exp/capacity-threshold-estimate/critical_capacity_at_observed_open_scale.csv"
SPATIAL_PATH = ROOT / "data/exp/capacity-threshold-estimate/spatial_scenario_threshold_summary.csv"
OUTPUT_PATH = ROOT / "data/results/tables/Table_aggregate_stress_load_and_reverse_capacity_thresholds.xlsx"
SHEET_NAME = "Aggregate Thresholds"

COLUMNS = [
    "Row Group",
    "Observation or Stress Scenario",
    "Hours Since Earthquake",
    "Observed Aggregate or Modeled Stress Load",
    "Opening Reference",
    "Capacity Threshold or Critical Capacity",
    "Aggregate Balance at Opening Reference",
    "Prefecture Arithmetic Minimum Openings",
    "Municipality-Contained Minimum Openings",
]

SCENARIO_LABELS = {
    "Observed-use stress, population weighted": "10,467-scaled population-weighted stress",
    "Observed-use stress, central housing-loss weighted": "10,467-scaled central-loss stress",
    "Observed-use stress, high housing-loss weighted": "10,467-scaled high-loss stress",
    "Modeled high housing-loss demand": "Modeled high housing-loss demand",
}


def build_table() -> pd.DataFrame:
    snapshots = pd.read_csv(SNAPSHOT_PATH)
    critical = pd.read_csv(CRITICAL_PATH)
    rows: list[dict[str, object]] = []

    for record in snapshots.itertuples(index=False):
        timestamp = pd.Timestamp(record[0]).strftime("%Y-%m-%d %H:%M JST")
        rows.append(
            {
                "Row Group": "Official snapshot x threshold",
                "Observation or Stress Scenario": timestamp,
                "Hours Since Earthquake": float(record[1]),
                "Observed Aggregate or Modeled Stress Load": int(round(float(record[2]))),
                "Opening Reference": int(record[3]),
                "Capacity Threshold or Critical Capacity": int(record[5]),
                "Aggregate Balance at Opening Reference": int(round(float(record[7]))),
                "Prefecture Arithmetic Minimum Openings": int(record[8]),
                "Municipality-Contained Minimum Openings": pd.NA,
            }
        )

    for record in critical.itertuples(index=False):
        demand = float(record[1])
        opening_reference = int(record[2])
        capacity = int(record[3])
        rows.append(
            {
                "Row Group": "Reverse capacity at 415 openings",
                "Observation or Stress Scenario": SCENARIO_LABELS[str(record[0])],
                "Hours Since Earthquake": pd.NA,
                "Observed Aggregate or Modeled Stress Load": int(round(demand)),
                "Opening Reference": opening_reference,
                "Capacity Threshold or Critical Capacity": capacity,
                "Aggregate Balance at Opening Reference": int(round(opening_reference * capacity - demand)),
                "Prefecture Arithmetic Minimum Openings": int(-(-int(round(demand)) // capacity)),
                "Municipality-Contained Minimum Openings": int(record[4]),
            }
        )

    table = pd.DataFrame(rows, columns=COLUMNS)
    if table.shape != (20, 9):
        raise RuntimeError(f"Expected a 20 x 9 table, found {table.shape}.")
    official = table["Row Group"].eq("Official snapshot x threshold")
    if int(official.sum()) != 16 or int((~official).sum()) != 4:
        raise RuntimeError("Expected 16 snapshot-threshold rows and four reverse-capacity rows.")
    peak_25 = table.loc[
        official
        & table["Observed Aggregate or Modeled Stress Load"].eq(10467)
        & table["Capacity Threshold or Critical Capacity"].eq(25)
    ].iloc[0]
    if int(peak_25["Aggregate Balance at Opening Reference"]) != -92 or int(peak_25["Prefecture Arithmetic Minimum Openings"]) != 419:
        raise RuntimeError("The peak 25-person aggregate calculation does not reconcile.")
    return table


def notes_table() -> pd.DataFrame:
    spatial = pd.read_csv(SPATIAL_PATH)
    high = spatial.loc[spatial["Demand Scenario"].eq("Observed-use stress, high housing-loss weighted")].set_index("Standardized Capacity per Shelter")
    if int(high.loc[25, "Minimum Open Shelters Required with Municipality Containment"]) != 441 or int(high.loc[50, "Minimum Open Shelters Required with Municipality Containment"]) != 235:
        raise RuntimeError("Municipality-contained comparison values do not reconcile.")
    return pd.DataFrame(
        [
            ("Official observations", "The four timestamps report prefecture-wide shelter users and open shelters. They are observed aggregates, not modeled local demand."),
            ("Stress-load scaling", "The available-snapshot maximum of 10,467 users scales three counterfactual residential demand surfaces; it is not an estimate of people refused shelter."),
            ("Opening reference", "The reverse-capacity rows use 415 openings, matching the reported open-shelter count at the 10,467-user snapshot."),
            ("Prefecture arithmetic", "Minimum openings equal ceil(total load / capacity), with no municipality boundaries or accessibility constraints."),
            ("Municipality containment", "Minimum openings equal the sum of municipality-specific ceilings and can exceed the prefecture arithmetic result."),
            ("25-person example", "For the 10,467-scaled high-loss stress surface, the prefecture arithmetic minimum is 419, whereas municipality containment requires 441."),
            ("50-person example", "For the same high-loss stress surface, the prefecture arithmetic minimum is 210, whereas municipality containment requires 235."),
            ("Reverse critical capacity", "The smallest integer capacity that keeps municipality-contained openings within 415 is a modeled requirement, not an observed facility capacity."),
            ("Capacity roles", "The 100-person threshold is central, 50 persons is a conservative stress case, and 25 and 200 persons are sensitivity cases."),
            ("Scope", "All rows precede network accessibility and facility-specific capacity constraints."),
        ],
        columns=["Note", "Definition or Limitation"],
    )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    thin = Border(bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9.5, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.5, color="172033")

    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:I{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 85
    worksheet.print_area = f"A1:I{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.22, right=0.22, top=0.30, bottom=0.30, header=0.12, footer=0.12)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 52
    group_fills = {"Official snapshot x threshold": "EAF2F8", "Reverse capacity at 415 openings": "FFF1D6"}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
        row[0].fill = PatternFill("solid", fgColor=group_fills[str(row[0].value)])
        row[0].font = Font(name="Aptos", size=8.5, bold=True, color="17365D")
        for index in range(2, 9):
            row[index].alignment = Alignment(horizontal="right", vertical="center")
        row[2].number_format = "0.0"
        for index in range(3, 9):
            row[index].number_format = "#,##0"
        if isinstance(row[6].value, (int, float)) and row[6].value < 0:
            row[6].font = Font(name="Aptos", size=8.5, bold=True, color="B42318")
            row[6].fill = PatternFill("solid", fgColor="FDE7E3")
        worksheet.row_dimensions[row[0].row].height = 34
    for column, width in {"A": 27, "B": 43, "C": 16, "D": 24, "E": 18, "F": 24, "G": 24, "H": 25, "I": 27}.items():
        worksheet.column_dimensions[column].width = width
    excel_table = Table(displayName="AggregateStressReverseCapacity", ref=f"A1:I{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(excel_table)

    notes = workbook["Notes"]
    notes.sheet_view.showGridLines = False
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 26
    notes.column_dimensions["B"].width = 110
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
        row[0].font = Font(name="Aptos", size=8.5, bold=True, color="17365D")
        notes.row_dimensions[row[0].row].height = 38
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 21 or worksheet.max_column != 9:
        raise RuntimeError(f"Unexpected main-sheet dimensions: {worksheet.max_row} x {worksheet.max_column}.")
    if worksheet.merged_cells.ranges:
        raise RuntimeError("Merged cells are not permitted in article-facing tables.")
    if workbook.sheetnames != [SHEET_NAME, "Notes"]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        build_table().to_excel(writer, sheet_name=SHEET_NAME, index=False)
        notes_table().to_excel(writer, sheet_name="Notes", index=False)
    style_workbook(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
