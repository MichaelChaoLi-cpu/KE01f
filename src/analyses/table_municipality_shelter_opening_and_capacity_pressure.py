#!/usr/bin/env python3
"""Municipality Shelter Opening and Capacity Pressure.

Plan: Report four prefecture summaries and the ten highest-pressure municipalities
within the single high-housing-loss-weighted stress scenario.
Framework: Sections 5-7 use municipality-contained opening requirements before
network assignment. The 50-person case is conservative and 100 persons is central.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
MUNICIPALITY_PATH = ROOT / "data/exp/capacity-threshold-estimate/municipality_capacity_thresholds.csv"
DEMAND_PATH = ROOT / "data/processed/kumamoto_prefecture_shelter_demand_preprocessed.parquet"
OUTPUT_PATH = ROOT / "data/results/tables/Table_municipality_shelter_opening_and_capacity_pressure.xlsx"
SHEET_NAME = "Municipality Pressure"
PRIMARY_SCENARIO = "Observed-use stress, high housing-loss weighted"

SCENARIO_LABELS = {
    "Modeled high housing-loss demand": "Modeled high housing-loss demand",
    "Observed-use stress, population weighted": "10,467-scaled population-weighted stress",
    "Observed-use stress, central housing-loss weighted": "10,467-scaled central-loss stress",
    "Observed-use stress, high housing-loss weighted": "10,467-scaled high-loss stress",
}

ENGLISH_NAMES = {
    "43100": "Kumamoto City", "43202": "Yatsushiro", "43204": "Arao",
    "43206": "Tamana", "43211": "Uto", "43213": "Uki", "43216": "Koshi",
    "43348": "Misato", "43404": "Kikuyo", "43443": "Mashiki",
    "43468": "Hikawa",
}

COLUMNS = [
    "Row Group",
    "Area",
    "Stress Scenario",
    "Stress Load",
    "General Shelters",
    "Maximum or Local Required Capacity (persons)",
    "Municipality-Contained Openings at 50 Persons",
    "Municipality-Contained Openings at 100 Persons",
    "Infeasible at 25 / 50 / 100 Persons",
    "Minimum Epicentral Distance (km)",
]


def build_full_results() -> pd.DataFrame:
    municipality = pd.read_csv(MUNICIPALITY_PATH, dtype={"Municipality Code": str})
    demand = pd.read_parquet(DEMAND_PATH, columns=["Municipality Code", "Epicentral Distance (km)"])
    demand["Municipality Code"] = demand["Municipality Code"].astype(str).str.zfill(5)
    distance = demand.groupby("Municipality Code", as_index=False)["Epicentral Distance (km)"].min()
    full = municipality.merge(distance, on="Municipality Code", how="left", validate="many_to_one")
    if full.shape[0] != 180 or full["Municipality Code"].nunique() != 45:
        raise RuntimeError("Expected 180 analytical rows covering 45 municipalities.")
    if full["Epicentral Distance (km)"].isna().any():
        raise RuntimeError("Municipality epicentral distance is missing.")
    return full


def build_table(full: pd.DataFrame) -> pd.DataFrame:
    scenario_order = [
        "Modeled high housing-loss demand",
        "Observed-use stress, population weighted",
        "Observed-use stress, central housing-loss weighted",
        "Observed-use stress, high housing-loss weighted",
    ]
    rows: list[dict[str, object]] = []
    for scenario in scenario_order:
        group = full.loc[full["Demand Scenario"].eq(scenario)]
        infeasible = [int((~group[f"Locally Feasible at {capacity} Persons"]).sum()) for capacity in (25, 50, 100)]
        rows.append(
            {
                "Row Group": "Prefecture summary",
                "Area": "Kumamoto Prefecture",
                "Stress Scenario": SCENARIO_LABELS[scenario],
                "Stress Load": int(round(group["Scenario Demand"].sum())),
                "General Shelters": int(group["General Shelters"].sum()),
                "Maximum or Local Required Capacity (persons)": float(group["Required Capacity if All General Shelters Open"].max()),
                "Municipality-Contained Openings at 50 Persons": int(group["Minimum Open Shelters Required at 50 Persons"].sum()),
                "Municipality-Contained Openings at 100 Persons": int(group["Minimum Open Shelters Required at 100 Persons"].sum()),
                "Infeasible at 25 / 50 / 100 Persons": " / ".join(map(str, infeasible)),
                "Minimum Epicentral Distance (km)": pd.NA,
            }
        )

    primary = full.loc[full["Demand Scenario"].eq(PRIMARY_SCENARIO)].copy()
    primary = primary.sort_values(
        ["Required Capacity if All General Shelters Open", "Municipality Code"],
        ascending=[False, True],
    ).head(10)
    for rank, record in enumerate(primary.itertuples(index=False), start=1):
        code = str(record[0]).zfill(5)
        name = ENGLISH_NAMES.get(code)
        if name is None:
            raise RuntimeError(f"Missing English municipality name for code {code}.")
        rows.append(
            {
                "Row Group": "High-loss pressure rank",
                "Area": f"{rank}. {name}",
                "Stress Scenario": "10,467-scaled high-loss stress",
                "Stress Load": int(round(float(record[5]))),
                "General Shelters": int(record[3]),
                "Maximum or Local Required Capacity (persons)": float(record[7]),
                "Municipality-Contained Openings at 50 Persons": int(record[10]),
                "Municipality-Contained Openings at 100 Persons": int(record[12]),
                "Infeasible at 25 / 50 / 100 Persons": " / ".join(
                    "No" if bool(value) else "Yes" for value in (record[9], record[11], record[13])
                ),
                "Minimum Epicentral Distance (km)": float(record[16]),
            }
        )

    table = pd.DataFrame(rows, columns=COLUMNS)
    if table.shape != (14, 10):
        raise RuntimeError(f"Expected a 14 x 10 table, found {table.shape}.")
    if table.iloc[4]["Area"] != "1. Yatsushiro" or table.iloc[4]["Stress Scenario"] != "10,467-scaled high-loss stress":
        raise RuntimeError("The within-scenario municipality ranking is inconsistent.")
    high_summary = table.loc[table["Stress Scenario"].eq("10,467-scaled high-loss stress")].iloc[0]
    if high_summary["Infeasible at 25 / 50 / 100 Persons"] != "3 / 0 / 0":
        raise RuntimeError("High-loss infeasible-municipality counts do not reconcile.")
    return table


def notes_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Ranking rule", "The ten municipality rows are ranked only within the 10,467-scaled high-loss stress scenario by demand divided by all local general shelters."),
            ("Observed-use boundary", "The 10,467 total is an observed prefecture aggregate used to scale a counterfactual residential stress surface; municipality values are not observed shelter use."),
            ("Required capacity", "Prefecture rows show the maximum municipality-level required average capacity; municipality rows show local stress load divided by all local general shelters."),
            ("Opening requirement", "Municipality-contained openings are summed after applying municipality-specific ceilings and cannot be compared directly with a single prefecture-wide ceiling."),
            ("Capacity roles", "The 100-person case is central and the 50-person case is a conservative stress case."),
            ("Feasibility field", "Prefecture rows report counts of infeasible municipalities; municipality rows report Yes or No in 25 / 50 / 100-person order."),
            ("High-loss infeasibility", "Under the 25-person sensitivity case, three municipalities are infeasible in the high-loss stress surface; none is infeasible at 50 or 100 persons."),
            ("Network scope", "Opening and reverse-capacity results precede network accessibility and facility-specific constraints."),
        ],
        columns=["Note", "Definition or Limitation"],
    )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    thin = Border(bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9.5, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.5, color="172033")

    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:J{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 82
    worksheet.print_area = f"A1:J{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.20, right=0.20, top=0.28, bottom=0.28, header=0.12, footer=0.12)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 58
    for row in worksheet.iter_rows(min_row=2):
        scope = str(row[0].value)
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
        row[0].fill = PatternFill("solid", fgColor="EAF2F8" if scope == "Prefecture summary" else "FFF1D6")
        row[0].font = Font(name="Aptos", size=8.5, bold=True, color="17365D")
        for index in (3, 4, 5, 6, 7, 9):
            row[index].alignment = Alignment(horizontal="right", vertical="center")
        for index in (3, 4, 6, 7):
            row[index].number_format = "#,##0"
        row[5].number_format = "0.0"
        row[9].number_format = "0.0"
        if "Yes" in str(row[8].value) or str(row[8].value).startswith(("1", "2", "3")):
            row[8].fill = PatternFill("solid", fgColor="FDE7E3")
        worksheet.row_dimensions[row[0].row].height = 38
    for column, width in {"A": 25, "B": 22, "C": 40, "D": 15, "E": 15, "F": 27, "G": 27, "H": 28, "I": 26, "J": 22}.items():
        worksheet.column_dimensions[column].width = width
    excel_table = Table(displayName="MunicipalityOpeningPressure", ref=f"A1:J{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(excel_table)

    notes = workbook["Notes"]
    notes.sheet_view.showGridLines = False
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 27
    notes.column_dimensions["B"].width = 110
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
        row[0].font = Font(name="Aptos", size=8.5, bold=True, color="17365D")
        notes.row_dimensions[row[0].row].height = 38
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 15 or worksheet.max_column != 10:
        raise RuntimeError(f"Unexpected main-sheet dimensions: {worksheet.max_row} x {worksheet.max_column}.")
    if worksheet.merged_cells.ranges:
        raise RuntimeError("Merged cells are not permitted in article-facing tables.")
    if workbook.sheetnames != [SHEET_NAME, "Notes"]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        build_table(build_full_results()).to_excel(writer, sheet_name=SHEET_NAME, index=False)
        notes_table().to_excel(writer, sheet_name="Notes", index=False)
    style_workbook(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
