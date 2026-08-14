#!/usr/bin/env python3
"""Network Accessibility and Robustness.

Plan: Provide a concise 26-row paper-facing synthesis of corrected accessibility,
shared-capacity allocation, matched facility loss, municipality mode gaps, and
screened single-shelter losses.
Framework: Sections 5-7 treat 10,467 as an aggregate-scaled stress surface and
report model explanation gaps rather than observed people refused shelter.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
ACCESS_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/walking_and_motorized_accessibility_summary.csv"
MIXED_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/mixed_mode_accessibility_summary.csv"
MUNICIPALITY_PATH = ROOT / "data/exp/prefecture-shelter-multimodal-access/municipality_mixed_mode_accessibility.csv"
CORE_PATH = ROOT / "data/exp/shared-capacity-multimodal-allocation/shared_capacity_mode_and_capacity_sensitivity_refined.csv"
OPENING_PATH = ROOT / "data/exp/shared-capacity-multimodal-allocation/matched_multimodal_opening_scale_sensitivity.csv"
FAILURE_PATH = ROOT / "data/exp/shared-capacity-multimodal-allocation/matched_multimodal_facility_unavailability.csv"
CRITICAL_PATH = ROOT / "data/exp/shelter-robustness/critical_single_shelter_loss.csv"
OUTPUT_PATH = ROOT / "data/results/tables/Table_network_accessibility_and_robustness.xlsx"
SHEET_NAME = "Network Robustness"

DEMAND_MEASURE = "Observed-Use Stress Demand High Housing-Loss Weighted"
TOTAL_STRESS_LOAD = 10467.0
ENGLISH_NAMES = {"43202": "Yatsushiro", "43100": "Kumamoto City"}
CRITICAL_NAMES = {
    "E4321300034111": "Toyofuku Elementary School Gymnasium (Uki)",
    "E4321300010111": "Ogawa Disaster Prevention Base Center (Uki)",
}

COLUMNS = [
    "Evidence Block",
    "Scenario",
    "Vehicle-Enabled Demand Share (%)",
    "Road Speed Factor",
    "Time Threshold (min)",
    "Capacity per Shelter",
    "Maximum Open Shelters",
    "Available Shelters",
    "Accessible or Assigned Demand",
    "Accessible or Assigned Share (%)",
    "Explanation Gap or Service Loss",
    "Solution Qualification",
]


def empty_row() -> dict[str, object]:
    return {column: pd.NA for column in COLUMNS}


def build_table() -> pd.DataFrame:
    access = pd.read_csv(ACCESS_PATH)
    mixed = pd.read_csv(MIXED_PATH)
    municipality = pd.read_csv(MUNICIPALITY_PATH, dtype={"Municipality Code": str})
    core = pd.read_csv(CORE_PATH)
    opening = pd.read_csv(OPENING_PATH)
    failure = pd.read_csv(FAILURE_PATH)
    critical = pd.read_csv(CRITICAL_PATH, dtype={"Shelter ID": str})
    rows: list[dict[str, object]] = []

    # Eight accessibility-bound rows: walking, three motorized speeds, and four mixed shares.
    walking = access.loc[
        access["Mode"].eq("Walking")
        & access["Demand Measure"].eq(DEMAND_MEASURE)
        & access["Time Threshold (min)"].eq(15)
    ].iloc[0]
    row = empty_row()
    row.update({
        "Evidence Block": "Accessibility bounds",
        "Scenario": "Walking bound",
        "Vehicle-Enabled Demand Share (%)": 0,
        "Time Threshold (min)": 15,
        "Accessible or Assigned Demand": round(float(walking["Accessible Demand"])),
        "Accessible or Assigned Share (%)": float(walking["Accessible Percent"]),
        "Explanation Gap or Service Loss": round(TOTAL_STRESS_LOAD - float(walking["Accessible Demand"])),
        "Solution Qualification": "4 km/h pedestrian-screened network",
    })
    rows.append(row)
    for speed_factor in (0.25, 0.50, 1.00):
        source = access.loc[
            access["Mode"].eq("Motorized")
            & access["Demand Measure"].eq(DEMAND_MEASURE)
            & access["Road Speed Factor"].eq(speed_factor)
            & access["Time Threshold (min)"].eq(15)
        ].iloc[0]
        row = empty_row()
        row.update({
            "Evidence Block": "Accessibility bounds",
            "Scenario": "Motorized bound",
            "Vehicle-Enabled Demand Share (%)": 100,
            "Road Speed Factor": speed_factor,
            "Time Threshold (min)": 15,
            "Accessible or Assigned Demand": round(float(source["Accessible Demand"])),
            "Accessible or Assigned Share (%)": float(source["Accessible Percent"]),
            "Explanation Gap or Service Loss": round(TOTAL_STRESS_LOAD - float(source["Accessible Demand"])),
            "Solution Qualification": "Every road-edge time scaled; connectors walked",
        })
        rows.append(row)
    for share in (0.25, 0.50, 0.75, 1.00):
        source = mixed.loc[mixed["Vehicle-Enabled Demand Share"].eq(share)].iloc[0]
        row = empty_row()
        row.update({
            "Evidence Block": "Accessibility bounds",
            "Scenario": "Mixed-mode accessibility",
            "Vehicle-Enabled Demand Share (%)": 100 * share,
            "Road Speed Factor": float(source["Motorized Road Speed Factor"]),
            "Time Threshold (min)": int(source["Time Threshold (min)"]),
            "Accessible or Assigned Demand": round(float(source["Accessible Demand"])),
            "Accessible or Assigned Share (%)": float(source["Accessible Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Explanation Gap"])),
            "Solution Qualification": "Capacity-free mode-share sensitivity",
        })
        rows.append(row)

    # Seven shared-capacity rows: five central-capacity mode shares plus matched 50-person cases.
    central = core.loc[core["Capacity per Open Shelter"].eq(100.0)].sort_values(
        "Vehicle-Enabled Demand Share"
    )
    for _, source in central.iterrows():
        share = float(source["Vehicle-Enabled Demand Share"])
        qualification = "Proven optimal" if bool(source["Proven Optimal"]) else f"Lower bound; solver gap {100 * float(source['MIP Gap']):.2f}%"
        row = empty_row()
        row.update({
            "Evidence Block": "Shared-capacity allocation",
            "Scenario": "Central 100-person capacity",
            "Vehicle-Enabled Demand Share (%)": 100 * share,
            "Road Speed Factor": 0.50 if share > 0 else pd.NA,
            "Time Threshold (min)": 15,
            "Capacity per Shelter": 100,
            "Maximum Open Shelters": 415,
            "Available Shelters": 1156,
            "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
            "Accessible or Assigned Share (%)": float(source["Served Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Model Explanation Gap"])),
            "Solution Qualification": qualification,
        })
        rows.append(row)
    for share in (0.0, 0.50):
        source = core.loc[
            core["Capacity per Open Shelter"].eq(50.0)
            & core["Vehicle-Enabled Demand Share"].eq(share)
        ].iloc[0]
        qualification = "Proven optimal" if bool(source["Proven Optimal"]) else f"Lower bound; solver gap {100 * float(source['MIP Gap']):.2f}%"
        row = empty_row()
        row.update({
            "Evidence Block": "Shared-capacity allocation",
            "Scenario": "Matched 50-person stress capacity",
            "Vehicle-Enabled Demand Share (%)": 100 * share,
            "Road Speed Factor": 0.50 if share > 0 else pd.NA,
            "Time Threshold (min)": 15,
            "Capacity per Shelter": 50,
            "Maximum Open Shelters": 415,
            "Available Shelters": 1156,
            "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
            "Accessible or Assigned Share (%)": float(source["Served Percent"]),
            "Explanation Gap or Service Loss": round(float(source["Model Explanation Gap"])),
            "Solution Qualification": qualification,
        })
        rows.append(row)

    # One matched opening-scale row under the central mixed-mode assumptions.
    source = opening.loc[
        opening["Opening Scenario"].eq("All 1,156 shelters selectable")
    ].iloc[0]
    qualification = "Proven optimal" if bool(source["Proven Optimal"]) else f"Lower bound; solver gap {100 * float(source['MIP Gap']):.2f}%"
    row = empty_row()
    row.update({
        "Evidence Block": "Shared-capacity allocation",
        "Scenario": "All general shelters selectable",
        "Vehicle-Enabled Demand Share (%)": 50,
        "Road Speed Factor": 0.50,
        "Time Threshold (min)": 15,
        "Capacity per Shelter": 100,
        "Maximum Open Shelters": 1156,
        "Available Shelters": 1156,
        "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
        "Accessible or Assigned Share (%)": float(source["Served Percent"]),
        "Explanation Gap or Service Loss": round(float(source["Model Explanation Gap"])),
        "Solution Qualification": qualification,
    })
    rows.append(row)

    # Seven matched facility-unavailability rows: baseline, three random means, three targeted cases.
    baseline = failure.loc[failure["Failure Mode"].eq("baseline")].iloc[0]
    row = empty_row()
    row.update({
        "Evidence Block": "Matched facility unavailability",
        "Scenario": "No-removal baseline",
        "Vehicle-Enabled Demand Share (%)": 50,
        "Road Speed Factor": 0.50,
        "Time Threshold (min)": 15,
        "Capacity per Shelter": 100,
        "Maximum Open Shelters": 415,
        "Available Shelters": int(baseline["Available Shelters"]),
        "Accessible or Assigned Demand": round(float(baseline["Maximum Served Demand"])),
        "Accessible or Assigned Share (%)": float(baseline["Served Percent"]),
        "Explanation Gap or Service Loss": round(float(baseline["Model Explanation Gap"])),
        "Solution Qualification": "Proven optimal matched baseline",
    })
    rows.append(row)
    for share in (0.10, 0.20, 0.30):
        source = failure.loc[
            failure["Failure Mode"].eq("random")
            & failure["Unavailability Share"].eq(share)
        ]
        mean_served = float(source["Maximum Served Demand"].mean())
        mean_share = float(source["Served Percent"].mean())
        max_gap = 100 * float(source["MIP Gap"].max())
        row = empty_row()
        row.update({
            "Evidence Block": "Matched facility unavailability",
            "Scenario": f"Random {int(100 * share)}% removal",
            "Vehicle-Enabled Demand Share (%)": 50,
            "Road Speed Factor": 0.50,
            "Time Threshold (min)": 15,
            "Capacity per Shelter": 100,
            "Maximum Open Shelters": 415,
            "Available Shelters": round(1156 * (1 - share)),
            "Accessible or Assigned Demand": round(mean_served),
            "Accessible or Assigned Share (%)": mean_share,
            "Explanation Gap or Service Loss": round(float(baseline["Maximum Served Demand"]) - mean_served),
            "Solution Qualification": f"Mean of 30 draws; maximum solver gap {max_gap:.2f}%",
        })
        rows.append(row)
    for share in (0.10, 0.20, 0.30):
        source = failure.loc[
            failure["Failure Mode"].eq("targeted_mixed_reachable_pressure")
            & failure["Unavailability Share"].eq(share)
        ].iloc[0]
        qualification = "Proven optimal" if bool(source["Proven Optimal"]) else f"Lower bound; solver gap {100 * float(source['MIP Gap']):.2f}%"
        row = empty_row()
        row.update({
            "Evidence Block": "Matched facility unavailability",
            "Scenario": f"Pressure-targeted {int(100 * share)}% removal",
            "Vehicle-Enabled Demand Share (%)": 50,
            "Road Speed Factor": 0.50,
            "Time Threshold (min)": 15,
            "Capacity per Shelter": 100,
            "Maximum Open Shelters": 415,
            "Available Shelters": int(source["Available Shelters"]),
            "Accessible or Assigned Demand": round(float(source["Maximum Served Demand"])),
            "Accessible or Assigned Share (%)": float(source["Served Percent"]),
            "Explanation Gap or Service Loss": round(float(baseline["Maximum Served Demand"]) - float(source["Maximum Served Demand"])),
            "Solution Qualification": qualification,
        })
        rows.append(row)

    # Two municipality rows; the gap column retains a gap, not a gain.
    walking_m = municipality.loc[municipality["Vehicle-Enabled Demand Share"].eq(0.0)].set_index("Municipality Code")
    motor_m = municipality.loc[municipality["Vehicle-Enabled Demand Share"].eq(1.0)].set_index("Municipality Code")
    comparison = walking_m[["Stress Load", "Accessible Demand", "Accessible Percent"]].join(
        motor_m[["Accessible Demand", "Accessible Percent"]],
        lsuffix=" Walk", rsuffix=" Motor",
    )
    comparison["Demand Gain"] = comparison["Accessible Demand Motor"] - comparison["Accessible Demand Walk"]
    for code in comparison.nlargest(2, "Demand Gain").index:
        source = comparison.loc[code]
        row = empty_row()
        row.update({
            "Evidence Block": "Municipality mode gap",
            "Scenario": ENGLISH_NAMES[str(code)],
            "Vehicle-Enabled Demand Share (%)": 100,
            "Road Speed Factor": 0.50,
            "Time Threshold (min)": 15,
            "Accessible or Assigned Demand": round(float(source["Accessible Demand Motor"])),
            "Accessible or Assigned Share (%)": float(source["Accessible Percent Motor"]),
            "Explanation Gap or Service Loss": round(float(source["Stress Load"] - source["Accessible Demand Motor"])),
            "Solution Qualification": f"Gain over walking bound: {float(source['Demand Gain']):.0f} persons",
        })
        rows.append(row)

    # Two conservative walking single-shelter rows.
    for _, source in critical.sort_values(
        "Single-Shelter Service-Loss Lower Bound", ascending=False
    ).head(2).iterrows():
        shelter_id = str(source["Shelter ID"])
        row = empty_row()
        row.update({
            "Evidence Block": "Single-shelter loss",
            "Scenario": CRITICAL_NAMES[shelter_id],
            "Vehicle-Enabled Demand Share (%)": 0,
            "Time Threshold (min)": 15,
            "Capacity per Shelter": 50,
            "Maximum Open Shelters": 415,
            "Available Shelters": 1155,
            "Accessible or Assigned Demand": round(float(source["Served Demand after Removal"])),
            "Accessible or Assigned Share (%)": 100 * float(source["Served Demand after Removal"]) / TOTAL_STRESS_LOAD,
            "Explanation Gap or Service Loss": round(float(source["Single-Shelter Service-Loss Lower Bound"])),
            "Solution Qualification": "Conservative walking screen; 30 shelters tested",
        })
        rows.append(row)

    table = pd.DataFrame(rows, columns=COLUMNS)
    if table.shape != (27, 12):
        raise RuntimeError(f"Expected a 27 x 12 table, found {table.shape}")
    if not table["Scenario"].eq("Random 20% removal").any():
        raise RuntimeError("Random 20-percent removal row is missing")
    if table["Scenario"].astype(str).str.contains(r"[^\x00-\x7F]", regex=True).any():
        raise RuntimeError("Non-ASCII scenario text found in the article-facing table")
    return table


def notes_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Demand surface", "All rows use the 10,467-scaled high-housing-loss stress surface. It is counterfactual residential demand, not observed municipality shelter use."),
            ("Accessibility bounds", "Walking is restrictive. Motorized times scale every road edge and retain walking connectors. Vehicle-enabled shares are sensitivity parameters."),
            ("Shared capacity", "Walking-only and vehicle-enabled components compete for the same shelter capacity. The central case allows at most 415 openings; the opening-scale row allows all 1,156 shelters to be selected."),
            ("Capacity roles", "The 100-person case is central; 50 persons is a conservative stress case."),
            ("Model explanation gap", "A gap is demand not assigned under the modeled rule set. It is not observed refusal or unsheltered population."),
            ("Random unavailability", "Rows report means from 30 reproducible draws. The qualification column reports the largest solver gap among the draws."),
            ("Targeted unavailability", "Pressure-targeted removal ranks shelters by 50% walking-reachable pressure plus 50% vehicle-enabled reachable pressure. It is an adverse test, not a probability forecast; nonzero solver gaps identify lower bounds."),
            ("Single-shelter screen", "The final two rows retain the conservative 50-person walking screen and are not an exhaustive ranking of all 1,156 shelters."),
            ("Primary comparison", "At a 50-percent vehicle-enabled demand share, the supported 50-to-100-person capacity gain is between zero and about 0.1 percentage point. Allowing all 1,156 shelters to be selected adds about 2.0 points, whereas changing mode availability adds about 20.1 points."),
        ],
        columns=["Note", "Definition or Limitation"],
    )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    thin = Border(bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=8.2, color="172033")
    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:L{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 75
    worksheet.print_area = f"A1:L{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.18, right=0.18, top=0.25, bottom=0.25, header=0.10, footer=0.10)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 58
    fills = {
        "Accessibility bounds": "DDEBF7",
        "Shared-capacity allocation": "FFF1D6",
        "Matched facility unavailability": "FDE7E3",
        "Municipality mode gap": "E8F3EC",
        "Single-shelter loss": "F2EBF6",
    }
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
        row[0].fill = PatternFill("solid", fgColor=fills[str(row[0].value)])
        row[0].font = Font(name="Aptos", size=8.2, bold=True, color="17365D")
        for index in range(2, 11):
            row[index].alignment = Alignment(horizontal="right", vertical="center")
        for index in (2, 4, 9):
            row[index].number_format = "0.0"
        row[3].number_format = "0.00"
        for index in (5, 6, 7, 8, 10):
            row[index].number_format = "#,##0"
        worksheet.row_dimensions[row[0].row].height = 37
    for column, width in {
        "A": 29, "B": 40, "C": 20, "D": 17, "E": 18, "F": 18,
        "G": 19, "H": 17, "I": 25, "J": 24, "K": 24, "L": 44,
    }.items():
        worksheet.column_dimensions[column].width = width
    excel_table = Table(displayName="NetworkAccessibilityRobustness", ref=f"A1:L{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(excel_table)

    notes = workbook["Notes"]
    notes.sheet_view.showGridLines = False
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 112
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
        row[0].font = Font(name="Aptos", size=8.2, bold=True, color="17365D")
        notes.row_dimensions[row[0].row].height = 40
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 28 or worksheet.max_column != 12:
        raise RuntimeError(f"Unexpected main-sheet dimensions: {worksheet.max_row} x {worksheet.max_column}")
    if worksheet.merged_cells.ranges:
        raise RuntimeError("Merged cells are not permitted")
    if workbook.sheetnames != [SHEET_NAME, "Notes"]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        build_table().to_excel(writer, sheet_name=SHEET_NAME, index=False)
        notes_table().to_excel(writer, sheet_name="Notes", index=False)
    style_workbook(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
