#!/usr/bin/env python3
"""Capacity Evidence and Threshold Calibration.

Plan: Summarize evidence coverage and calibrate standardized capacity cases.
Framework: Sections 5-7 distinguish general from welfare-specific supply and use
118 documented general shelters in three municipalities for plausibility checks.
The 100-person case is central; 50 persons is a conservative stress case.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
AUDIT_PATH = ROOT / "data/exp/prefecture-shelter-capacity-audit/prefecture_shelter_capacity_evidence_audit.parquet"
TIER_PATH = ROOT / "data/exp/prefecture-shelter-capacity-audit/capacity_evidence_tier_summary.csv"
THRESHOLD_PATH = ROOT / "data/exp/capacity-threshold-estimate/official_capacity_threshold_calibration.csv"
OUTPUT_PATH = ROOT / "data/results/tables/Table_capacity_evidence_and_threshold_calibration.xlsx"
SHEET_NAME = "Capacity Calibration"

COLUMNS = [
    "Row Group",
    "Evidence Tier or Calibration Item",
    "Shelter Service Class",
    "Shelters",
    "Inventory Share (%)",
    "Official Numeric Capacities",
    "Official Capacity Median (persons)",
    "Sample at or above Threshold (%)",
]

TIER_LABELS = {
    "A_numeric_capacity": "A - Official numeric capacity",
    "B_gymnasium_area": "B - Documented gymnasium area",
    "C_gross_facility_area": "C - Gross facility area",
    "D_school_identity_only": "D - School identity only",
    "E_public_facility_identity_only": "E - Public-facility identity only",
    "F_no_capacity_source_link": "F - No capacity-source link",
}


def build_table() -> pd.DataFrame:
    audit = pd.read_parquet(AUDIT_PATH)
    tier_summary = pd.read_csv(TIER_PATH)
    threshold = pd.read_csv(THRESHOLD_PATH)
    tier_crosstab = pd.crosstab(audit["Capacity Evidence Tier"], audit["Shelter Service Class"])

    rows: list[dict[str, object]] = []
    for record in tier_summary.itertuples(index=False):
        tier = str(record[0])
        general = int(tier_crosstab.loc[tier].get("general", 0))
        welfare = int(tier_crosstab.loc[tier].get("welfare_specific", 0))
        numeric = audit.loc[
            audit["Capacity Evidence Tier"].eq(tier), "Official Numeric Capacity"
        ].dropna()
        rows.append(
            {
                "Row Group": "Evidence tier",
                "Evidence Tier or Calibration Item": TIER_LABELS[tier],
                "Shelter Service Class": f"General {general:,}; welfare-specific {welfare:,}",
                "Shelters": int(record.Shelters),
                "Inventory Share (%)": float(record.Percent),
                "Official Numeric Capacities": int(len(numeric)),
                "Official Capacity Median (persons)": numeric.median(),
                "Sample at or above Threshold (%)": pd.NA,
            }
        )

    for service_class, label in (
        ("general", "General-shelter inventory"),
        ("welfare_specific", "Welfare-specific inventory"),
    ):
        subset = audit.loc[audit["Shelter Service Class"].eq(service_class)]
        numeric = subset["Official Numeric Capacity"].dropna()
        rows.append(
            {
                "Row Group": "Service class",
                "Evidence Tier or Calibration Item": label,
                "Shelter Service Class": "General" if service_class == "general" else "Welfare-specific",
                "Shelters": int(len(subset)),
                "Inventory Share (%)": 100.0,
                "Official Numeric Capacities": int(len(numeric)),
                "Official Capacity Median (persons)": numeric.median(),
                "Sample at or above Threshold (%)": pd.NA,
            }
        )

    for record in threshold.itertuples(index=False):
        capacity = int(record[0])
        role = "Stress case" if capacity == 50 else "Central case" if capacity == 100 else "Sensitivity case"
        rows.append(
            {
                "Row Group": "Threshold calibration",
                "Evidence Tier or Calibration Item": f"{capacity} persons - {role}",
                "Shelter Service Class": "General",
                "Shelters": int(record[1]),
                "Inventory Share (%)": 100.0,
                "Official Numeric Capacities": int(record[1]),
                "Official Capacity Median (persons)": float(record[3]),
                "Sample at or above Threshold (%)": float(record[7]),
            }
        )

    table = pd.DataFrame(rows, columns=COLUMNS)
    if table.shape != (12, 8):
        raise RuntimeError(f"Expected a 12 x 8 table, found {table.shape}.")
    if int(table.loc[table["Row Group"].eq("Evidence tier"), "Shelters"].sum()) != len(audit):
        raise RuntimeError("Evidence-tier counts do not reconcile to the master inventory.")
    general = table.loc[table["Evidence Tier or Calibration Item"].eq("General-shelter inventory")].iloc[0]
    if int(general["Shelters"]) != 1156 or int(general["Official Numeric Capacities"]) != 118:
        raise RuntimeError("General-shelter inventory or calibration sample is inconsistent.")
    return table


def notes_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Purpose", "Evidence tiers describe source strength; they are not capacity estimates."),
            ("Service classes", "Welfare-specific shelters are excluded from unrestricted general-population supply."),
            ("Calibration sample", "Official numeric capacities cover 118 general shelters in Uki, Uto, and Yatsushiro; the sample is not prefecture-representative."),
            ("Central case", "A standardized capacity of 100 persons per general shelter is the central analytical case."),
            ("Stress case", "A standardized capacity of 50 persons per general shelter is retained as a conservative stress case."),
            ("Sensitivity cases", "The 25- and 200-person thresholds bound the capacity sensitivity analysis."),
            ("Area evidence", "Gymnasium and gross facility areas are not converted to people because usable shelter-floor shares are unavailable."),
        ],
        columns=["Note", "Definition or Limitation"],
    )


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    thin = Border(bottom=Side(style="thin", color="D0D5DD"))
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=9, color="172033")

    worksheet = workbook[SHEET_NAME]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:H{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 90
    worksheet.print_area = f"A1:H{worksheet.max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 44
    group_fills = {"Evidence tier": "EAF2F8", "Service class": "E8F3EC", "Threshold calibration": "FFF1D6"}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
        row[0].fill = PatternFill("solid", fgColor=group_fills[str(row[0].value)])
        row[0].font = Font(name="Aptos", size=9, bold=True, color="17365D")
        for index in range(3, 8):
            row[index].alignment = Alignment(horizontal="right", vertical="center")
        row[3].number_format = "#,##0"
        row[4].number_format = "0.0"
        row[5].number_format = "#,##0"
        row[6].number_format = "#,##0"
        row[7].number_format = "0.0"
        worksheet.row_dimensions[row[0].row].height = 32
    for column, width in {"A": 21, "B": 34, "C": 30, "D": 12, "E": 18, "F": 21, "G": 24, "H": 25}.items():
        worksheet.column_dimensions[column].width = width
    excel_table = Table(displayName="CapacityEvidenceCalibration", ref=f"A1:H{worksheet.max_row}")
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(excel_table)

    notes = workbook["Notes"]
    notes.sheet_view.showGridLines = False
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 105
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    notes.row_dimensions[1].height = 30
    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
        row[0].font = Font(name="Aptos", size=9, bold=True, color="17365D")
        notes.row_dimensions[row[0].row].height = 34
    workbook.save(path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[SHEET_NAME]
    if worksheet.max_row != 13 or worksheet.max_column != 8:
        raise RuntimeError(f"Unexpected main-sheet dimensions: {worksheet.max_row} x {worksheet.max_column}.")
    if worksheet.merged_cells.ranges:
        raise RuntimeError("Merged cells are not permitted in article-facing tables.")
    if workbook.sheetnames != [SHEET_NAME, "Notes"]:
        raise RuntimeError(f"Unexpected sheet order: {workbook.sheetnames}")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                    raise RuntimeError(f"Formula error text found in {sheet.title}!{cell.coordinate}.")


def main() -> None:
    table = build_table()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        notes_table().to_excel(writer, sheet_name="Notes", index=False)
    style_workbook(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
